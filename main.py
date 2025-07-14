from ultralytics import YOLO
import cv2
import numpy as np
import ast
import re
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
import string
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import zipfile


def safe_math_eval(expr):
    """Вычисление математических выражений"""
    try:
        def eval_frac(match):
            try:
                return str(float(match.group(1)) / float(match.group(2)))
            except:
                return match.group(0)

        evaluated = re.sub(r'(\d+\.?\d*)\s*/\s*(\d+\.?\d*)', eval_frac, expr)
        return ast.literal_eval(evaluated)
    except:
        return expr


def get_settings(file_path):
    """Получение настроек из конфиг-файла"""
    with open(file_path, 'r', encoding='utf8') as file:
        lines = [line.strip() for line in file if line.strip()]
        param_lines = lines[-4:]

        params = []
        for line in param_lines:
            if ';' not in line:
                continue

            value = line.split(';', 1)[1].strip()
            if line.startswith('Параметры матрицы камеры'):
                try:
                    evaluated = safe_math_eval(value)
                    params.append(evaluated)
                except:
                    params.append(value)
            else:
                try:
                    params.append(ast.literal_eval(value))
                except (ValueError, SyntaxError):
                    params.append(value)

        return params


def get_img(frame):
    """Получение обработанного изображения"""
    global settings
    return cv2.undistort(frame, camera_matrix, distortion_coefficients)


def detect_objects(image, model_path='runs/detect/train/weights/best.pt', bool=False):
    """Обнаружение объектов на изображении"""
    global settings, count
    model = YOLO(model_path)
    results = model(image)
    annotated_frame = results[0].plot()

    for result in results:
        for box in result.boxes:
            info = get_info(settings[3], settings[2], settings[0], count, box, result)
            cv2.putText(
                annotated_frame,
                str(info[2]),
                info[4],
                cv2.FONT_HERSHEY_SIMPLEX,
                0.5,
                (78, 200, 255),
                2
            )

            if bool:
                count += 1
                crop(info, image, count)
                report2exel(info)
                print(info)

    cv2.imshow("YOLO Detection", annotated_frame)


def cords_transform(img_cords, cam_pose, h, camera_matrix):
    """Трансформация координат"""
    cx, cy = cam_pose
    ix, iy = img_cords

    fx = camera_matrix[0][0]
    fy = camera_matrix[1][1]
    CX = camera_matrix[0][2]
    CY = camera_matrix[1][2]

    X = cx + (ix - CX) / fx * h
    Y = cy - (iy - CY) / fy * h

    return (round(X, 3), round(Y, 3))


def get_info(cam_pose, h, camera_matrix, count, box, result):
    """Получение информации об объекте"""
    x, y, _, _ = map(int, box.xywh[0])
    x1, y1, x2, y2 = map(int, box.xyxy[0])
    class_id = int(box.cls[0])
    confidence = float(box.conf[0])
    class_name = result.names[class_id]
    file_name = f'obj_n{count}.jpg'
    time = datetime.now().strftime("%Y-%m-%d,%H:%M:%S")

    return [
        class_id,
        class_name,
        cords_transform((x, y), cam_pose, h, camera_matrix),
        ((x1, y1), (x2, y2)),
        (x, y),
        round(confidence, 2),
        time,
        file_name
    ]


def crop(info, img, count):
    """Обрезка и сохранение изображения распознанного объекта"""
    y1, y2 = info[3][0][1], info[3][1][1]
    x1, x2 = info[3][0][0], info[3][1][0]
    res = img[y1:y2, x1:x2]
    cv2.imwrite(f'Q:/ITMOPROJECT/for_script/img2excel/obj_n{count}.jpg', res)


def load_counter():
    """Загрузка значения счетчика нумераций изображений из файла"""
    try:
        with open('for_script/counter.txt', 'r') as f:
            return int(f.read())
    except (FileNotFoundError, ValueError):
        return 0


def save_counter(count):
    """Сохранение значения счетчика нумераций изображений в файл"""
    os.makedirs('for_script', exist_ok=True)
    with open('for_script/counter.txt', 'w') as f:
        f.write(str(count))


def report2exel(info):
    """Сохранение данных в Excel файл"""
    global count

    if info[5] < 0.5:
        fill_color = PatternFill(
            start_color="FFCCCC",
            end_color="FFCCCC",
            fill_type="solid"
        )
    elif 0.5 <= info[5] <= 0.75:
        fill_color = PatternFill(
            start_color="FFFFCC",
            end_color="FFFFCC",
            fill_type="solid"
        )
    else:
        fill_color = PatternFill(
            start_color="CCFFCC",
            end_color="CCFFCC",
            fill_type="solid"
        )

    try:
        wb = load_workbook('for_script/report.xlsx')
        ws = wb.active
    except (FileNotFoundError, KeyError, zipfile.BadZipFile):
        wb = Workbook()
        ws = wb.active
        count = 0

        headers = [
            "ID класса",
            "Имя класса",
            "Трансформированные координаты (в метрах)",
            "Координаты BBox",
            "Координаты объекта относительно изображения (в px)",
            "Уверенность",
            "Дата и время",
            "Имя файла изображения распознанного объекта"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    row_num = ws.max_row + 1 if ws.max_row > 1 else 2

    for col_idx in range(8):
        col_letter = get_column_letter(col_idx + 1)
        cell = ws[f"{col_letter}{row_num}"]
        value = info[col_idx]

        if isinstance(value, (tuple, list)):
            cell.value = ', '.join(map(str, value))
        else:
            cell.value = str(value)

        cell.alignment = Alignment(horizontal='center')

        if col_letter == 'F':
            cell.fill = fill_color

    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col),
            default=10
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    wb.save('for_script/report.xlsx')


def main():
    """Основная функция программы"""
    global settings, count, save_requested, camera_matrix, distortion_coefficients

    cap = cv2.VideoCapture(1)
    count = load_counter()
    save_requested = False

    settings = get_settings('for_script/settings.txt')
    camera_matrix = np.array(settings[0])
    distortion_coefficients = np.array(settings[1])

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        key = cv2.waitKey(1)
        if key == ord(' '):
            save_requested = True
        elif key == ord('q'):
            save_counter(count)
            break

        if save_requested:
            detect_objects(get_img(frame), bool=True)
            save_requested = False
        else:
            detect_objects(get_img(frame))

    cap.release()
    cv2.destroyAllWindows()


if __name__ == "__main__":
    main()